"""
Hawaii Election Dashboard - data preprocessor.

Reads the three source Excel files and emits a single `data.js` consumed by
`dashboard.html`. All cleaning, joining, and category mapping happens here so
the front-end is pure render + filter.

Run: python preprocess.py
"""
from __future__ import annotations

import json
import re
from pathlib import Path
import warnings

import pandas as pd

warnings.filterwarnings("ignore")

HERE = Path(__file__).parent
ELECTION_XLSX = HERE / "Hawaii_Elections_Combined_l (1).xlsx"
MAPPING_XLSX = HERE / "Precint_Mapping.xlsx"
DEMOGRAPHICS_XLSX = HERE / "cleaned_precincts.xlsx"
ACS_DISTRICT_XLSX = HERE / "acs_district_demographics.xlsx"
OUT_JS = HERE / "data.js"

KEEP_YEARS = {2022, 2024}


# --- race label cleanup ---

PRESIDENT_PREFIXES = [
    "U.S. President and Vice President",
    "U.S. President",
    "President and Vice President",
    "President",
    "For PRESIDENT",
    "For President",
]


def clean_race_label(race: str) -> str:
    if not isinstance(race, str):
        return race
    race = race.strip()
    if "PRESIDENT" not in race.upper():
        return race
    party_match = re.search(r"\(([A-Z]+)\)", race)
    party = party_match.group(1) if party_match else ""
    candidate_name = re.sub(r"\([A-Z]+\)\s*", "", race).strip()
    for prefix in PRESIDENT_PREFIXES:
        candidate_name = re.compile(re.escape(prefix), re.IGNORECASE).sub("", candidate_name).strip()
    if candidate_name and len(candidate_name) > 2:
        return f"U.S. President ({party}) {candidate_name}" if party else f"U.S. President - {candidate_name}"
    return race


# --- precinct helpers ---

PRECINCT_PREFIX_RE = re.compile(r"^(MAIL|IN-PERSON)\s+", flags=re.IGNORECASE)


def split_precinct(p: str) -> tuple[str, str]:
    """Return (base_precinct, vote_type) where vote_type is 'M', 'I', or ''."""
    if not isinstance(p, str):
        return str(p), ""
    p = p.strip()
    m = PRECINCT_PREFIX_RE.match(p)
    if not m:
        return p, ""
    kind = m.group(1).upper()
    base = PRECINCT_PREFIX_RE.sub("", p)
    return base, "M" if kind == "MAIL" else "I"


# --- mapping ---


def load_mapping() -> dict[str, dict]:
    df = pd.read_excel(MAPPING_XLSX, dtype={"PRECINCT": str})
    df.columns = df.columns.str.strip()
    df["PRECINCT"] = df["PRECINCT"].astype(str).str.strip()
    out = {}
    for _, row in df.iterrows():
        p = row["PRECINCT"]
        def clean(v):
            if pd.isna(v):
                return None
            try:
                return str(int(float(v)))
            except (TypeError, ValueError):
                return str(v).strip()
        out[p] = {
            "hd": clean(row.get("STATEREP")),
            "sd": clean(row.get("STATESENATE")),
            "cd": clean(row.get("U.S.REP")),
            "cc": clean(row.get("COUNCIL")),
        }
    return out


# --- election results ---


def load_election(mapping: dict[str, dict]) -> tuple[list[dict], list[str]]:
    df = pd.read_excel(ELECTION_XLSX, sheet_name="Election Results", dtype={"Precinct": str})
    df.columns = df.columns.str.strip()
    # normalize column names
    df = df.rename(columns={
        "year": "Year", "Election Type": "Election Type", "Precinct": "Precinct",
        "race": "Race", "candidate": "Candidate", "votes": "Votes", "percentage": "Percentage",
    })
    df = df[df["Year"].isin(KEEP_YEARS)].copy()
    df["Precinct"] = df["Precinct"].astype(str).str.strip()

    # split precinct → base + vote type
    bases, vts = zip(*df["Precinct"].map(split_precinct))
    df["base"] = list(bases)
    df["vt"] = list(vts)

    # parse votes / drop zero
    df["Votes"] = pd.to_numeric(df["Votes"], errors="coerce").fillna(0).astype(int)
    df = df[df["Votes"] > 0].copy()

    # clean race labels (and trim everything after '/')
    df["Race"] = df["Race"].astype(str).str.split("/").str[0].str.strip().map(clean_race_label)

    # join district info via mapping
    rows = []
    for _, r in df.iterrows():
        m = mapping.get(r["base"], {})
        rows.append({
            "y": int(r["Year"]),
            "t": r["Election Type"],
            "p": r["base"],
            "vt": r["vt"],
            "r": r["Race"],
            "c": r["Candidate"],
            "v": int(r["Votes"]),
            "hd": m.get("hd"),
            "sd": m.get("sd"),
            "cd": m.get("cd"),
            "cc": m.get("cc"),
        })

    races = sorted(df["Race"].dropna().unique().tolist())
    return rows, races


# --- turnout ---


def load_turnout(mapping: dict[str, dict]) -> list[dict]:
    df = pd.read_excel(ELECTION_XLSX, sheet_name="Turnout Statistics", dtype={"Precinct": str})
    df.columns = df.columns.str.strip()
    df = df.rename(columns={"year": "Year"})
    df = df[df["Year"].isin(KEEP_YEARS)].copy()
    df["Precinct"] = df["Precinct"].astype(str).str.strip()
    df["Precinct"] = df["Precinct"].str.replace(PRECINCT_PREFIX_RE, "", regex=True)

    party_cols = [c for c in [
        "libertarianParty", "greenParty", "democraticParty", "nonpartisan",
        "republicanParty", "alohaAinaParty", "constitutionParty",
        "americanShoppingParty", "specialOnly", "solidarityParty",
        "liberationParty", "weThePeople", "noLabels", "noPartySelected",
        "multiPartyVoting", "independent",
    ] if c in df.columns]

    agg = {"precinctRegistration": "first", "mailTurnout": "sum"}
    for c in party_cols:
        agg[c] = "sum"

    df = df.groupby(["Year", "Election Type", "Precinct"], as_index=False).agg(agg)
    df["turnoutPercent"] = (df["mailTurnout"] / df["precinctRegistration"] * 100).round(1)

    # drop primary rows where both major parties are zero (artifact of source)
    if "democraticParty" in df.columns and "republicanParty" in df.columns:
        df = df[
            (df["Election Type"] == "General")
            | (df["democraticParty"] != 0)
            | (df["republicanParty"] != 0)
        ].copy()

    out = []
    for _, r in df.iterrows():
        m = mapping.get(r["Precinct"], {})
        row = {
            "y": int(r["Year"]),
            "t": r["Election Type"],
            "p": r["Precinct"],
            "reg": int(r["precinctRegistration"]) if pd.notna(r["precinctRegistration"]) else 0,
            "tn": int(r["mailTurnout"]) if pd.notna(r["mailTurnout"]) else 0,
            "tp": float(r["turnoutPercent"]) if pd.notna(r["turnoutPercent"]) else 0.0,
            "hd": m.get("hd"),
            "sd": m.get("sd"),
            "cd": m.get("cd"),
            "cc": m.get("cc"),
        }
        for c in party_cols:
            row[c] = int(r[c]) if pd.notna(r[c]) else 0
        out.append(row)
    return out


# --- demographics ---

# Census-column categorization. Returns dict[category, list[col]] plus the per-category
# `total` reference column used for % normalization.
CATEGORY_TOTAL_COL = {
    "Housing Occupancy": "Total:",
    "Housing Tenure": "Total:_1",
    "Householder Race": "Total:_2",
    "Household Size": "Total:_3",
    "Age of Householder": "Total:_4",
    "Presence of Children (Owner)": "Owner occupied:_1",
    "Presence of Children (Renter)": "Renter occupied:_1",
    "Household Type": "Total:_6",
}


def categorize_demographics(columns: list[str]) -> dict[str, list[str]]:
    cats: dict[str, list[str]] = {
        "Housing Occupancy": [],
        "Housing Tenure": [],
        "Householder Race": [],
        "Household Size": [],
        "Age of Householder (Owner)": [],
        "Age of Householder (Renter)": [],
        "Presence of Children (Owner)": [],
        "Presence of Children (Renter)": [],
        "Household Type": [],
    }
    for col in columns:
        if col == "Precinct":
            continue
        lo = col.lower()
        if col.startswith("Owner occupied:") or col.startswith("Renter occupied:"):
            continue
        if col == "Total:" or col == "Occupied" or col == "Vacant":
            cats["Housing Occupancy"].append(col)
        elif "owned with a mortgage" in lo or "owned free and clear" in lo or (col == "Renter occupied" and "householder" not in lo):
            cats["Housing Tenure"].append(col)
        elif "householder who is" in lo:
            cats["Householder Race"].append(col)
        elif "-person household" in lo or "-or-more-person household" in lo:
            cats["Household Size"].append(col)
        elif "householder" in lo and "years" in lo and "children" not in lo:
            if col.endswith(".1"):
                cats["Age of Householder (Renter)"].append(col)
            else:
                cats["Age of Householder (Owner)"].append(col)
        elif "children under 18" in lo:
            if col.endswith(".1"):
                cats["Presence of Children (Renter)"].append(col)
            else:
                cats["Presence of Children (Owner)"].append(col)
        elif any(k in lo for k in [
            "family households", "nonfamily households", "married couple",
            "male householder", "female householder", "householder living alone",
            "householder not living alone",
        ]):
            cats["Household Type"].append(col)
    return {k: v for k, v in cats.items() if v}


# 55-59 + 60-64 collapse to 55-64 in pyramid view (matches existing behavior)
AGE_COLLAPSE = {
    "Householder 15 to 24 years": "15 to 24",
    "Householder 25 to 34 years": "25 to 34",
    "Householder 35 to 44 years": "35 to 44",
    "Householder 45 to 54 years": "45 to 54",
    "Householder 55 to 59 years": "55 to 64",
    "Householder 60 to 64 years": "55 to 64",
    "Householder 65 to 74 years": "65 to 74",
    "Householder 75 to 84 years": "75 to 84",
    "Householder 85 years and over": "85+",
}

CHILDREN_LABEL = {
    "With children under 18 years": "With Children",
    "No children under 18 years": "No Children",
}


def load_demographics() -> dict:
    df = pd.read_excel(DEMOGRAPHICS_XLSX, dtype={"Precinct": str})
    df.columns = df.columns.str.strip()
    df["Precinct"] = df["Precinct"].astype(str).str.strip()

    cats = categorize_demographics(list(df.columns))

    # per precinct, build {category: {label: value}, totals: {...}}
    out: dict[str, dict] = {}
    for _, row in df.iterrows():
        p = row["Precinct"]
        rec: dict = {"totals": {}, "categories": {}}
        # totals per category
        for cat, total_col in CATEGORY_TOTAL_COL.items():
            if total_col in df.columns:
                v = row.get(total_col)
                rec["totals"][cat] = float(v) if pd.notna(v) else 0.0

        # standard categories (non-pyramid)
        for cat_name in ["Housing Occupancy", "Housing Tenure", "Householder Race", "Household Size", "Household Type"]:
            if cat_name not in cats:
                continue
            data = {}
            for col in cats[cat_name]:
                if col.startswith("Total:") or col.startswith("Owner occupied:") or col.startswith("Renter occupied:"):
                    continue
                v = row.get(col)
                if pd.notna(v):
                    data[col] = float(v)
            if data:
                rec["categories"][cat_name] = data

        # age pyramid (owner + renter combined into single 6-bucket structure)
        owner_age = {}
        renter_age = {}
        for col in cats.get("Age of Householder (Owner)", []):
            base_col = col.replace(".1", "")
            label = AGE_COLLAPSE.get(base_col)
            if label is None:
                continue
            v = row.get(col)
            if pd.notna(v):
                owner_age[label] = owner_age.get(label, 0.0) + float(v)
        for col in cats.get("Age of Householder (Renter)", []):
            base_col = col.replace(".1", "")
            label = AGE_COLLAPSE.get(base_col)
            if label is None:
                continue
            v = row.get(col)
            if pd.notna(v):
                renter_age[label] = renter_age.get(label, 0.0) + float(v)
        if owner_age or renter_age:
            rec["age_pyramid"] = {"owner": owner_age, "renter": renter_age}

        # children pyramid
        owner_kids = {}
        renter_kids = {}
        for col in cats.get("Presence of Children (Owner)", []):
            base_col = col.replace(".1", "")
            label = CHILDREN_LABEL.get(base_col)
            if label is None:
                continue
            v = row.get(col)
            if pd.notna(v):
                owner_kids[label] = float(v)
        for col in cats.get("Presence of Children (Renter)", []):
            base_col = col.replace(".1", "")
            label = CHILDREN_LABEL.get(base_col)
            if label is None:
                continue
            v = row.get(col)
            if pd.notna(v):
                renter_kids[label] = float(v)
        if owner_kids or renter_kids:
            rec["children_pyramid"] = {"owner": owner_kids, "renter": renter_kids}

        out[p] = rec
    return out


_ACS_SHEET_RE = re.compile(
    r"^State (Representative|Senator),\s*Dist\s+(\d+)$", re.IGNORECASE
)

# Canonical income bucket labels (source labels vary slightly across sheets).
_INCOME_NORMALIZE = [
    (re.compile(r"0\s*to\s*\$?50[,.]?000", re.I), "$0 - $50K"),
    (re.compile(r"50[,.]?000.*to.*100[,.]?000", re.I), "$50K - $100K"),
    (re.compile(r"100[,.]?000.*to.*150[,.]?000", re.I), "$100K - $150K"),
    (re.compile(r"150[,.]?000.*to.*200[,.]?000", re.I), "$150K - $200K"),
    (re.compile(r"200[,.]?000\s*\+", re.I), "$200K+"),
]
_INCOME_ORDER = ["$0 - $50K", "$50K - $100K", "$100K - $150K", "$150K - $200K", "$200K+"]


def _normalize_income_label(raw: str) -> str | None:
    for pat, canon in _INCOME_NORMALIZE:
        if pat.search(raw):
            return canon
    return None


def _parse_section(ws, header_label: str) -> list[tuple[str, float, int]]:
    """Find a section header (e.g. 'Age', 'Ethnicity', 'Income') anywhere in the
    sheet and walk the rows below it, reading (label, value) from the header's
    column and the next column. Returns a list of (clean_label, value, indent)
    triples preserving the source-sheet order. Indent = whitespace level
    (0 = top race, 1 = sub-race) derived from leading NBSP/space count."""
    header_row = header_col = None
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip().lower() == header_label.lower():
                header_row, header_col = r, c
                break
        if header_row is not None:
            break
    if header_row is None:
        return []

    rows: list[tuple[str, float, int]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        label = ws.cell(r, header_col).value
        value = ws.cell(r, header_col + 1).value
        if label is None and value is None:
            break
        if not isinstance(label, str) or value is None:
            continue
        # Indent = number of leading whitespace chars / 4 (NBSP and regular both count)
        lead = len(label) - len(label.lstrip(" \xa0"))
        indent = 1 if lead >= 14 else 0  # 12 sp = top-level, 16 sp = sub
        clean = label.strip(" \xa0")
        rows.append((clean, float(value), indent))
    return rows


def load_acs_district() -> dict[str, dict]:
    """Read per-district ACS demographics from `acs_district_demographics.xlsx`.

    Returns {district_label: {age: [[label, value, indent], ...],
                              ethnicity: [...],
                              income: [...]}}
    where district_label is "HD-N" / "SD-N".
    """
    if not ACS_DISTRICT_XLSX.exists():
        print(f"  (no {ACS_DISTRICT_XLSX.name} — skipping ACS district data)")
        return {}
    import openpyxl  # local import; only needed here
    wb = openpyxl.load_workbook(ACS_DISTRICT_XLSX, data_only=True)
    out: dict[str, dict] = {}
    for sheet_name in wb.sheetnames:
        m = _ACS_SHEET_RE.match(sheet_name.strip())
        if not m:
            continue
        kind, num = m.group(1).lower(), m.group(2)
        prefix = "HD" if kind == "representative" else "SD"
        label = f"{prefix}-{int(num)}"

        ws = wb[sheet_name]
        age_rows = _parse_section(ws, "Age")
        eth_rows = _parse_section(ws, "Ethnicity")
        inc_rows_raw = _parse_section(ws, "Income")

        # Normalize income labels to canonical set + order
        income_map = {}
        for raw_label, val, _ in inc_rows_raw:
            canon = _normalize_income_label(raw_label)
            if canon:
                income_map[canon] = val
        inc_rows = [(c, income_map[c], 0) for c in _INCOME_ORDER if c in income_map]

        # Normalize age labels: strip stray whitespace ('  85+' on some sheets)
        age_rows = [(lbl.strip(), v, 0) for (lbl, v, _) in age_rows]

        if not (age_rows or eth_rows or inc_rows):
            continue
        out[label] = {"age": age_rows, "ethnicity": eth_rows, "income": inc_rows}
    return out


def main():
    print("Loading mapping...")
    mapping = load_mapping()
    print(f"  {len(mapping)} precincts mapped")

    print("Loading election results...")
    election_rows, races = load_election(mapping)
    print(f"  {len(election_rows):,} non-zero rows across {len(races)} races")

    print("Loading turnout...")
    turnout_rows = load_turnout(mapping)
    print(f"  {len(turnout_rows):,} turnout rows")

    print("Loading demographics...")
    demographics = load_demographics()
    print(f"  {len(demographics)} precincts with demographic data")

    print("Loading ACS district demographics...")
    acs_district = load_acs_district()
    print(f"  {len(acs_district)} districts with ACS data")

    # build district lists from actual data
    def collect(field):
        vals = set()
        for r in election_rows:
            v = r.get(field)
            if v is not None:
                vals.add(v)
        def sort_key(x):
            try:
                return (0, int(x))
            except ValueError:
                return (1, x)
        return sorted(vals, key=sort_key)

    # ---- index-encode strings to shrink JSON ~7x ----
    def build_dict(items: list[str]) -> tuple[list[str], dict[str, int]]:
        uniq = sorted({x for x in items if x is not None})
        return uniq, {v: i for i, v in enumerate(uniq)}

    races_list, race_ix = build_dict([r["r"] for r in election_rows])
    cands_list, cand_ix = build_dict([r["c"] for r in election_rows])
    precincts_list, prec_ix = build_dict(
        [r["p"] for r in election_rows] + [r["p"] for r in turnout_rows]
    )
    types_list, type_ix = build_dict([r["t"] for r in election_rows])
    years_list = sorted({r["y"] for r in election_rows})
    year_ix = {y: i for i, y in enumerate(years_list)}

    # columnar arrays for election: [yi, ti, pi, vt, ri, ci, v]
    e_yi, e_ti, e_pi, e_vt, e_ri, e_ci, e_v = [], [], [], [], [], [], []
    for r in election_rows:
        e_yi.append(year_ix[r["y"]])
        e_ti.append(type_ix[r["t"]])
        e_pi.append(prec_ix[r["p"]])
        e_vt.append(r["vt"])  # 'M', 'I', or ''
        e_ri.append(race_ix[r["r"]])
        e_ci.append(cand_ix[r["c"]])
        e_v.append(r["v"])

    # columnar turnout
    party_keys = sorted({
        k for r in turnout_rows for k in r.keys()
        if k not in {"y", "t", "p", "reg", "tn", "tp", "hd", "sd", "cd", "cc"}
    })
    t_yi, t_ti, t_pi, t_reg, t_tn, t_tp = [], [], [], [], [], []
    t_party = {k: [] for k in party_keys}
    for r in turnout_rows:
        t_yi.append(year_ix[r["y"]])
        t_ti.append(type_ix[r["t"]])
        t_pi.append(prec_ix[r["p"]])
        t_reg.append(r["reg"])
        t_tn.append(r["tn"])
        t_tp.append(r["tp"])
        for k in party_keys:
            t_party[k].append(r.get(k, 0))

    # district lookup keyed by precinct index
    mapping_by_pi: dict[int, dict] = {}
    for p, m in mapping.items():
        if p in prec_ix:
            mapping_by_pi[prec_ix[p]] = m

    # demographics keyed by precinct index
    demo_by_pi: dict[int, dict] = {}
    for p, rec in demographics.items():
        if p in prec_ix:
            demo_by_pi[prec_ix[p]] = rec

    bundle = {
        "years": years_list,
        "types": types_list,
        "precincts": precincts_list,
        "races": races_list,
        "candidates": cands_list,
        "election": {
            "yi": e_yi, "ti": e_ti, "pi": e_pi, "vt": e_vt,
            "ri": e_ri, "ci": e_ci, "v": e_v,
        },
        "turnout": {
            "yi": t_yi, "ti": t_ti, "pi": t_pi,
            "reg": t_reg, "tn": t_tn, "tp": t_tp,
            "parties": t_party,
        },
        "demographics": demo_by_pi,
        "acs_district": acs_district,
        "mapping": mapping_by_pi,
        "districts": {
            "hd": collect("hd"),
            "sd": collect("sd"),
            "cd": collect("cd"),
            "cc": collect("cc"),
        },
        "generated": pd.Timestamp.now().isoformat(),
    }

    js = "window.__DATA = " + json.dumps(bundle, separators=(",", ":")) + ";\n"
    OUT_JS.write_text(js, encoding="utf-8")
    print(f"\nWrote {OUT_JS} ({OUT_JS.stat().st_size / 1024:.1f} KB)")


if __name__ == "__main__":
    main()
